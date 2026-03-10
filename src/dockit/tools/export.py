# export.py - Excel 导出
"""开庭记录表、期限列表导出为 Excel"""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from ..db.db import get_connection, init_db

logger = logging.getLogger(__name__)


def export_court_sessions(archive_dir: str | Path, output_path: str | Path | None = None) -> Path:
    """
    导出开庭记录表到 Excel。

    Returns:
        输出文件路径
    """
    archive_dir = Path(archive_dir).expanduser()
    init_db(archive_dir)
    if output_path is None:
        output_path = archive_dir / "开庭记录表.xlsx"
    else:
        output_path = Path(output_path)

    conn = get_connection(archive_dir)
    rows = conn.execute("""
        SELECT c.case_number, c.cause_of_action, c.court_name, c.plaintiff, c.defendant,
               s.session_time, s.location, s.judge, s.session_result
        FROM court_sessions s
        JOIN cases c ON s.case_id = c.id
        ORDER BY s.session_time
    """).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "开庭记录"

    headers = ["案号", "案由", "法院", "原告", "被告", "开庭时间", "地点", "审判长", "状态"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(output_path)
    logger.info("已导出: %s", output_path)
    return output_path


def export_deadlines(archive_dir: str | Path, output_path: str | Path | None = None) -> Path:
    """
    导出期限列表到 Excel。

    Returns:
        输出文件路径
    """
    archive_dir = Path(archive_dir).expanduser()
    init_db(archive_dir)
    if output_path is None:
        output_path = archive_dir / "关键期限.xlsx"
    else:
        output_path = Path(output_path)

    conn = get_connection(archive_dir)
    rows = conn.execute("""
        SELECT c.case_number, c.cause_of_action, d.deadline_type, d.due_date, d.is_completed, d.notes
        FROM deadlines d
        JOIN cases c ON d.case_id = c.id
        ORDER BY d.due_date
    """).fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "关键期限"

    headers = ["案号", "案由", "期限类型", "到期日", "是否完成", "备注"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            v = val
            if col_idx == 5:
                v = "是" if val else "否"
            ws.cell(row=row_idx, column=col_idx, value=v)

    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    wb.save(output_path)
    logger.info("已导出: %s", output_path)
    return output_path


if __name__ == "__main__":  # pragma: no cover
    import os
    import yaml

    logging.basicConfig(level=logging.INFO)
    from ..config_path import get_config_path
    config_path = get_config_path()
    with open(config_path, encoding="utf-8") as f:
        config = yaml.safe_load(f)
    raw = config["archive_dir"].replace("$HOME", os.path.expanduser("~"))
    archive_dir = str(Path(raw).expanduser())
    export_court_sessions(archive_dir)
    export_deadlines(archive_dir)
    print("导出完成:", archive_dir)
